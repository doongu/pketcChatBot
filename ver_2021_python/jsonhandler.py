import crawler
from bs4 import BeautifulSoup
import requests
import urllib3
import json
import time

from openpyxl import load_workbook, Workbook
import sys, os, time


title_list, date_list, url_list = crawler.ce_notice_crawler("https://cms.pknu.ac.kr/ced/main.do")
def ce_notice():
    global title_list, date_list, url_list
    data = {
  "version": "2.0",
  "template": {
    "outputs": [
      {
        "listCard": {
          "header": {
            "title": "학과 공지사항입니다."
          },
          "items": [
            {
              "title": title_list[0],
              "description": date_list[0],
              
              "link": {
                "web": url_list[0]
              }
            },
            {
              "title": title_list[1],
              "description": date_list[1],
              
              "link": {
                "web": url_list[1]
              }
            },
            {
              "title": title_list[2],
              "description": date_list[2],
              
              "link": {
                "web": url_list[2]
              }
            },
            {
              "title": title_list[3],
              "description": date_list[3],
              
              "link": {
                "web": url_list[3]
              }
            },
            {
              "title": title_list[4],
              "description": date_list[4],
              
              "link": {
                "web": url_list[4]
              }
            }
          ],
          "buttons": [
            {
              "label": "학과 홈페이지",
              "action": "webLink",
              "webLinkUrl": "https://cms.pknu.ac.kr/ced/main.do"
            }
          ]
        }
      }
    ]
  }
}
    return data

def ohora(excel_path):
    load_wb = load_workbook(excel_path, data_only=True)
    load_ws = load_wb['Sheet1']

    url_list = []
    i = 3
    while True:
        temp = load_ws.cell(i, 9).value
        if temp:
            url_list.append( temp )
        else:
            break
        i += 1
    
    return_answer = ""
    for i in range(len(url_list)):
        url1 = url_list[i]   # 품절
        url_idx = url1.find("product_no")
        answer = ""
        for i in range(url_idx, len(url1)):
            if url1[i].isdigit():
                answer += url1[i]
            if url1[i] == "&":
                break

        select_url = url1
        headers={"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"}
        request = requests.post(select_url,headers = headers, verify=False)
        soup = BeautifulSoup(request.text, "html.parser")
        # soup = str(soup)\
        answer = soup.find("div", {"class":"xans-element- xans-product xans-product-action "})
        answer2 = answer.find_all("a")
        if "displaynone" in str(answer2[0]):
            if i == len(excel_path) - 1:
                  return_answer += "OOOO"
                  break
            return_answer += "OOOO\n"
        else:
            if i == len(excel_path) - 1:
                return_answer += "XXXX"
                break
            return_answer += "XXXX\n"


        return_str ={ "version": "2.0",
        "template": {
          "outputs": [
            {
              "simpleText": {
                "text": return_answer
              }
            }
          ]
        }
      }
        return return_str

#	chrome_options = webdriver.ChromeOptions()

#	chrome_options.add_argument('--headless')

#	chrome_options.add_argument('--no-sandbox')

#	chrome_options.add_argument('--disable-dev-shm-usage')

#    browser.set_window_position(0, 0)
#    browser.set_window_size(2000, 2000)


    # write_wb =  Workbook()
    # write_ws = write_wb.active
    #
    # for i in range(0, len(input_list)):
    #     write_ws.cell(i+1, 5, input_list[i])
    # write_wb.save(save_path + "/ohara_result.xlsx")
    # write_wb.close()
